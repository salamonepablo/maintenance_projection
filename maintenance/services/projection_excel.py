"""
Exportador de proyecciones a Excel con formato condicional.
Replica el estilo y colores del código VB Materfer.
"""
from __future__ import annotations

from datetime import date
from typing import TYPE_CHECKING

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

if TYPE_CHECKING:
    from .projection_grid import ModuleProjectionRow


class ProjectionExcelExporter:
    """
    Exporta proyecciones a Excel con formato profesional.
    
    Formato condicional (umbrales):
    - DA (1.500.000 km): Rojo suave cuando excede
    - P (750.000 km): Amarillo suave cuando excede  
    - BI (375.000 km): Naranja suave cuando excede
    - A (187.500 km): Celeste claro cuando excede
    """
    
    # Colores para formato condicional (del VB)
    COLORS = {
        "DA": {
            "bg": "FFE0E0",  # Rosa claro
            "text": "FF0000"  # Rojo
        },
        "P": {
            "bg": "FFFFE0",  # Amarillo suave
            "text": "808000"  # Amarillo oscuro
        },
        "BI": {
            "bg": "FFE0B0",  # Naranja suave
            "text": "FF8000"  # Naranja
        },
        "A": {
            "bg": "E0F0FF",  # Celeste claro
            "text": "0000FF"  # Azul
        },
        "reset": {
            "bg": "D8D8D8",  # Gris para reseteos
            "text": "000000"  # Negro
        }
    }
    
    def export(
        self,
        projections: dict[int, list[ModuleProjectionRow]],
        filepath: str,
        monthly_km: int = 12_500
    ) -> None:
        """
        Exporta proyecciones a archivo Excel.
        
        Args:
            projections: Dict con proyecciones por módulo
            filepath: Ruta del archivo a crear
            monthly_km: Km promedio mensual usado en proyección
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Proyección Mantenimiento"
        
        # Configuración de página
        self._setup_sheet(ws)
        
        # Información de parámetros
        current_row = self._add_header(ws, monthly_km)
        
        # Encabezados de columnas
        current_row = self._add_column_headers(ws, current_row, projections)
        
        # Datos de proyección por módulo
        for module_id in sorted(projections.keys()):
            rows = projections[module_id]
            current_row = self._add_module_data(ws, current_row, rows)
        
        # Ajustar anchos de columna
        self._adjust_column_widths(ws)
        
        # Guardar archivo
        wb.save(filepath)
    
    def _setup_sheet(self, ws) -> None:
        """Configuración inicial de la hoja"""
        # Configurar orientación y márgenes
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
    
    def _add_header(self, ws, monthly_km: int) -> int:
        """Agrega header con información de parámetros"""
        # Título
        ws['A1'] = 'PROYECCIÓN DE MANTENIMIENTO - FLOTA CSR'
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A1:F1')
        
        # Fecha de generación
        ws['A2'] = f'Fecha de generación: {date.today().strftime("%d/%m/%Y")}'
        ws['A2'].font = Font(size=10)
        
        # Km promedio mensual
        ws['A3'] = f'Km promedio mensual: {monthly_km:,} km'
        ws['A3'].font = Font(size=10, color='0000FF')  # Azul para inputs
        
        return 5  # Fila donde empezar los datos
    
    def _add_column_headers(
        self,
        ws,
        start_row: int,
        projections: dict[int, list[ModuleProjectionRow]]
    ) -> int:
        """Agrega encabezados de columnas"""
        # Obtener cantidad de meses de la primera proyección
        first_module = next(iter(projections.values()))
        num_months = len(first_module[0].cells)
        
        headers = [
            'N° Módulo',
            'Intervención',
            'Fecha Último Evento',
            'Km Acumulado'
        ]
        
        # Agregar headers de meses
        for cell in first_module[0].cells:
            headers.append(cell.month_date.strftime('%b %y'))
        
        # Escribir headers
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=start_row, column=col, value=header)
            cell.font = Font(bold=True, size=10)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
            
            # Bordes
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            cell.border = thin_border
        
        return start_row + 1
    
    def _add_module_data(
        self,
        ws,
        start_row: int,
        rows: list[ModuleProjectionRow]
    ) -> int:
        """Agrega datos de un módulo (4 filas: DA, P, BI, A)"""
        current_row = start_row
        
        for row in rows:
            # Columna 1: N° Módulo
            cell = ws.cell(row=current_row, column=1, value=row.module_id)
            cell.font = Font(bold=True, size=10)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Columna 2: Tipo de Intervención
            cell = ws.cell(row=current_row, column=2, value=row.intervention_type)
            cell.font = Font(bold=True, size=10)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Color del tipo según jerarquía
            color = self.COLORS[row.intervention_type]["text"]
            cell.font = Font(bold=True, size=10, color=color)
            
            # Columna 3: Fecha último evento
            date_str = (
                row.last_event_date.strftime('%d/%m/%Y') 
                if row.last_event_date 
                else 'N/A'
            )
            cell = ws.cell(row=current_row, column=3, value=date_str)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Columna 4: Km inicial acumulado
            cell = ws.cell(row=current_row, column=4, value=row.initial_km)
            cell.number_format = '#,##0'
            cell.alignment = Alignment(horizontal='right', vertical='center')
            
            # Aplicar formato condicional
            self._apply_cell_format(
                cell,
                row.initial_km,
                row.intervention_type,
                False
            )
            
            # Columnas 5+: Proyección mensual
            for idx, grid_cell in enumerate(row.cells, start=5):
                cell = ws.cell(row=current_row, column=idx)
                
                # Si hay reseteo, mostrar código de intervención
                if grid_cell.is_reset_point:
                    cell.value = grid_cell.intervention_code
                    cell.font = Font(bold=True, size=10)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    # Color gris para reseteos
                    cell.fill = PatternFill(
                        start_color=self.COLORS["reset"]["bg"],
                        end_color=self.COLORS["reset"]["bg"],
                        fill_type='solid'
                    )
                else:
                    # Mostrar km acumulado
                    cell.value = grid_cell.km_accumulated
                    cell.number_format = '#,##0'
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                    
                    # Aplicar formato condicional
                    self._apply_cell_format(
                        cell,
                        grid_cell.km_accumulated,
                        row.intervention_type,
                        grid_cell.exceeds_threshold
                    )
                
                # Bordes
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                cell.border = thin_border
            
            current_row += 1
        
        # Espacio entre módulos
        return current_row + 1
    
    def _apply_cell_format(
        self,
        cell,
        km_value: int,
        intervention_type: str,
        exceeds_threshold: bool
    ) -> None:
        """Aplica formato condicional a una celda según su valor"""
        if exceeds_threshold:
            colors = self.COLORS[intervention_type]
            cell.fill = PatternFill(
                start_color=colors["bg"],
                end_color=colors["bg"],
                fill_type='solid'
            )
            cell.font = Font(color=colors["text"], size=10)
    
    def _adjust_column_widths(self, ws) -> None:
        """Ajusta ancho de columnas automáticamente"""
        column_widths = {
            1: 12,  # N° Módulo
            2: 12,  # Intervención
            3: 18,  # Fecha
            4: 15,  # Km Acumulado
        }
        
        # Columnas de meses (ancho estándar)
        for col in range(5, ws.max_column + 1):
            column_widths[col] = 12
        
        for col, width in column_widths.items():
            ws.column_dimensions[get_column_letter(col)].width = width

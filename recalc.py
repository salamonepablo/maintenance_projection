#!/usr/bin/env python3
"""
Script para recalcular fórmulas de Excel usando LibreOffice.
Adaptado del skill de xlsx.

Uso:
    python recalc.py archivo.xlsx [timeout_seconds]

Ejemplo:
    python recalc.py output.xlsx 30
"""
import sys
import os
import json
import subprocess
import time
from pathlib import Path


def setup_libreoffice_macro():
    """Configura macro de LibreOffice para recalcular fórmulas"""
    # Determinar directorio de macros según OS
    if sys.platform == 'darwin':  # macOS
        macro_dir = Path.home() / 'Library/Application Support/LibreOffice/4/user/Scripts/python'
    else:  # Linux
        macro_dir = Path.home() / '.config/libreoffice/4/user/Scripts/python'
    
    macro_dir.mkdir(parents=True, exist_ok=True)
    
    macro_content = '''
import uno
from com.sun.star.beans import PropertyValue

def recalculate():
    """Recalcula todas las fórmulas en el documento"""
    desktop = XSCRIPTCONTEXT.getDesktop()
    model = desktop.getCurrentComponent()
    
    if model:
        sheets = model.getSheets()
        for i in range(sheets.getCount()):
            sheet = sheets.getByIndex(i)
            sheet.calculateAll()
    
    return None
'''
    
    macro_file = macro_dir / 'recalc_macro.py'
    macro_file.write_text(macro_content)
    
    return macro_file


def recalculate_excel(filepath: str, timeout: int = 30) -> dict:
    """
    Recalcula fórmulas de Excel usando LibreOffice.
    
    Args:
        filepath: Ruta al archivo Excel
        timeout: Timeout en segundos
        
    Returns:
        Dict con status, total_errors, error_summary
    """
    filepath = os.path.abspath(filepath)
    
    if not os.path.exists(filepath):
        return {
            'status': 'error',
            'message': f'Archivo no encontrado: {filepath}'
        }
    
    # Verificar LibreOffice
    libreoffice_cmd = None
    if sys.platform == 'darwin':  # macOS
        libreoffice_cmd = '/Applications/LibreOffice.app/Contents/MacOS/soffice'
    else:  # Linux
        for cmd in ['libreoffice', 'soffice']:
            if subprocess.run(['which', cmd], capture_output=True).returncode == 0:
                libreoffice_cmd = cmd
                break
    
    if not libreoffice_cmd or not os.path.exists(libreoffice_cmd):
        return {
            'status': 'error',
            'message': 'LibreOffice no está instalado. Instala: sudo apt install libreoffice'
        }
    
    try:
        # Abrir archivo en headless mode y recalcular
        cmd = [
            libreoffice_cmd,
            '--headless',
            '--calc',
            '--convert-to', 'xlsx',
            '--outdir', os.path.dirname(filepath),
            filepath
        ]
        
        result = subprocess.run(
            cmd,
            timeout=timeout,
            capture_output=True,
            text=True
        )
        
        if result.returncode != 0:
            return {
                'status': 'error',
                'message': f'Error al recalcular: {result.stderr}'
            }
        
        # Verificar errores en fórmulas usando openpyxl
        errors = check_formula_errors(filepath)
        
        if errors['total_errors'] > 0:
            return {
                'status': 'errors_found',
                'total_errors': errors['total_errors'],
                'total_formulas': errors['total_formulas'],
                'error_summary': errors['error_summary']
            }
        
        return {
            'status': 'success',
            'total_errors': 0,
            'total_formulas': errors['total_formulas']
        }
    
    except subprocess.TimeoutExpired:
        return {
            'status': 'error',
            'message': f'Timeout después de {timeout} segundos'
        }
    except Exception as e:
        return {
            'status': 'error',
            'message': str(e)
        }


def check_formula_errors(filepath: str) -> dict:
    """
    Verifica errores en fórmulas de Excel.
    
    Returns:
        Dict con total_errors, total_formulas, error_summary
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        return {
            'total_errors': 0,
            'total_formulas': 0,
            'error_summary': {}
        }
    
    wb = load_workbook(filepath, data_only=False)
    
    error_types = ['#REF!', '#DIV/0!', '#VALUE!', '#N/A', '#NAME?', '#NUM!', '#NULL!']
    error_summary = {}
    total_errors = 0
    total_formulas = 0
    
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        
        for row in sheet.iter_rows():
            for cell in row:
                # Verificar si es una fórmula
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    total_formulas += 1
                
                # Verificar si tiene error
                if cell.value in error_types:
                    error_type = cell.value
                    cell_ref = f'{sheet_name}!{cell.coordinate}'
                    
                    if error_type not in error_summary:
                        error_summary[error_type] = {
                            'count': 0,
                            'locations': []
                        }
                    
                    error_summary[error_type]['count'] += 1
                    error_summary[error_type]['locations'].append(cell_ref)
                    total_errors += 1
    
    return {
        'total_errors': total_errors,
        'total_formulas': total_formulas,
        'error_summary': error_summary
    }


def main():
    if len(sys.argv) < 2:
        print('Uso: python recalc.py <archivo.xlsx> [timeout_seconds]')
        sys.exit(1)
    
    filepath = sys.argv[1]
    timeout = int(sys.argv[2]) if len(sys.argv) > 2 else 30
    
    print(f'Recalculando fórmulas en: {filepath}')
    print(f'Timeout: {timeout} segundos')
    print()
    
    result = recalculate_excel(filepath, timeout)
    
    # Imprimir resultado como JSON
    print(json.dumps(result, indent=2))
    
    # Exit code según resultado
    if result['status'] == 'success':
        sys.exit(0)
    elif result['status'] == 'errors_found':
        sys.exit(1)
    else:
        sys.exit(2)


if __name__ == '__main__':
    main()
